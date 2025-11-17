from .utils.contexto_dinamico import gerar_contexto_dinamico
import re
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, Count
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Paciente, Consulta, Dentista, Procedimento, Assinatura, Pagamento, Income, Expense
from .forms import PacienteForm, ProcedimentoForm, IncomeForm, ExpenseForm
from .forms_consulta import ConsultaForm
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.utils.dateparse import parse_datetime
from django.db.models import Count, Sum, Avg
import calendar
from datetime import timedelta
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .decorators import require_active_subscription
from django.conf import settings
from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from openai import OpenAI
import json
import os
import uuid
import mercadopago
from decimal import Decimal
from django.urls import reverse
import openpyxl
from openpyxl.utils import get_column_letter



# inicializa o client globalmente mas de forma segura
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

def _odontoia_system_prompt(user):
    from .utils.contexto_dinamico import gerar_contexto_dinamico

    nome = user.first_name or user.username
    contexto_dinamico = gerar_contexto_dinamico(user)

    return f"""
Você é o **Assistente Oficial do OdontoIA**, um sistema de gestão odontológica voltado a dentistas, clínicas e secretárias.

Regras fundamentais:
- Sempre responda em **português (pt-BR)**, de forma profissional, breve e empática.
- **Nunca forneça informações clínicas** sobre pacientes, diagnósticos ou tratamentos odontológicos.
- O foco é **ensinar o usuário a usar o sistema OdontoIA**.
- O público-alvo são **usuários do sistema (dentistas, secretárias, administradores)**, não pacientes.
- Se o usuário perguntar “como faço X”, explique o caminho pelo sistema (menus e páginas) de forma prática.
- Use os nomes exatos dos menus do OdontoIA: **Dashboard**, **Pacientes**, **Consultas**, **Procedimentos**, **Agenda**, **Financeiro**, **/admin** (apenas administradores).

Contexto real do sistema:
- O cadastro de **dentistas** é feito **somente por usuários administradores**, no painel **/admin**, usando o campo **CRO**.
- O cadastro de **pacientes** e **procedimentos** é feito diretamente no app.
- O módulo de **consultas** permite marcar, editar e concluir atendimentos.
- O módulo de **financeiro** mostra comissões, faturamento e lucro líquido.
- O sistema possui **teste gratuito de 7 dias**, após o qual a conta precisa de assinatura ativa.
- Se o usuário perguntar sobre **planos ou preços**, explique que há planos **Básico**, **Profissional** e **Premium**.
- Se o trial estiver próximo de expirar, lembre da página de planos (https://odontoia.codertec.com.br).

Comportamento:
- Seja proativo e claro (ex: “Vá até ‘Pacientes’ > ‘Novo Paciente’”).
- Se a ação for restrita (como cadastrar dentista), avise que só **administradores** podem fazer.
- Se não souber, oriente a procurar o suporte da Codertec.

📊 **Dados em tempo real do sistema:**
{contexto_dinamico}

Contexto atual:
- Usuário logado: {nome}
- Data e hora: {timezone.now().strftime('%d/%m/%Y %H:%M')}
    """.strip()


def _get_openai_client():
    try:
        from openai import OpenAI
    except Exception as e:
        return None, f"SDK OpenAI não disponível: {e}"

    # tenta .env → settings
    api_key = os.getenv("OPENAI_API_KEY") or getattr(
        settings, "OPENAI_API_KEY", None)
    if not api_key:
        return None, "OPENAI_API_KEY não configurada (adicione no .env ou nas variáveis do serviço)."

    try:
        client = OpenAI(api_key=api_key)  # sem proxies/kwargs estranhos
        return client, None
    except Exception as e:
        return None, f"Falha ao criar cliente OpenAI: {e}"

# ── Endpoint de diagnóstico (opcional, ajuda no debug) ─────────────────────

@csrf_exempt
@login_required
@require_GET
def chat_diag(request):
    has_key = bool(os.getenv("OPENAI_API_KEY") or getattr(
        settings, "OPENAI_API_KEY", None))
    return JsonResponse({
        "ok": True,
        "user": request.user.username,
        "openai_key_present": has_key,
    })


# ── Endpoint principal do chat ──────────────────────────────────────────────
@csrf_exempt                 # evita dor de cabeça com CSRF no fetch
@login_required              # exige login (se 302 → login)
@require_POST
def chat_odontoia_api(request):
    # lê payload
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
        user_msg = (payload.get("message") or "").strip()
    except Exception:
        return JsonResponse({"error": "Payload inválido."}, status=400)

    if not user_msg:
        return JsonResponse({"error": "Mensagem vazia."}, status=400)

    # cliente OpenAI
    client, err = _get_openai_client()
    if err:
        # devolve o motivo real pro front (facilita corrigir config)
        return JsonResponse({"error": err}, status=500)

    # histórico (curto) na sessão
    history = request.session.get("chat_history", [])
    messages = [
        {"role": "system", "content": _odontoia_system_prompt(request.user)}]
    messages.extend(history[-20:])  # últimos turnos
    messages.append({"role": "user", "content": user_msg})

    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            temperature=0.2,
        )
        answer = (resp.choices[0].message.content or "").strip()
    except Exception as e:
        # log no servidor e motivo no front
        print("OpenAI error:", repr(e))
        return JsonResponse({"error": f"OpenAI falhou: {e}"}, status=502)

    # salva histórico
    history.append({"role": "user", "content": user_msg})
    history.append({"role": "assistant", "content": answer})
    request.session["chat_history"] = history

    return JsonResponse({"answer": answer})


# === LOGIN ===
def user_login(request):
    if request.user.is_authenticated:
        return redirect('clinic:dashboard')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)

                # 🔄 limpa mensagens anteriores
                storage = messages.get_messages(request)
                for _ in storage:
                    pass

                messages.success(request, f"Bem-vindo(a), {user.username}!")

                # Se ainda não viu a tela de boas vindas
                if not request.session.get('onboarding_done', False):
                    return redirect('clinic:onboarding')

                # Caso contrário vai para dashboard
                return redirect('clinic:dashboard')

            else:
                messages.error(request, "Usuário ou senha inválidos.")
        else:
            messages.error(request, "Erro ao validar o formulário.")
    else:
        form = AuthenticationForm()

    return render(request, 'clinic/login.html', {'form': form})


# === LOGOUT ===
def user_logout(request):
    """
    Faz logout, limpa mensagens antigas e redireciona para login sem conflito.
    """
    logout(request)

    # ⚙️ Limpa mensagens antigas antes de adicionar uma nova
    storage = messages.get_messages(request)
    for _ in storage:
        pass  # força limpeza

    messages.success(request, "Você saiu do sistema com sucesso.")
    return redirect("clinic:login")


# permite AJAX sem token no protótipo (depois ajustamos segurança)
@csrf_exempt
@login_required
@require_active_subscription
def consulta_create_ajax(request):
    if request.method == "POST":
        try:
            consulta = Consulta.objects.create(
                paciente_id=request.POST.get("paciente"),
                dentista_id=request.POST.get("dentista"),
                procedimento_id=request.POST.get("procedimento"),
                data=parse_datetime(request.POST.get("data")),
                observacoes=request.POST.get("observacoes", ""),
                owner=request.user  # 🔥 ESSENCIAL
            )
            return JsonResponse({"success": True, "id": consulta.id})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método inválido"})


@login_required
@require_active_subscription
def consultas_list(request):
    # Obter filtros via GET
    search = request.GET.get('search')
    status = request.GET.get('status')
    data_filtro = request.GET.get('data')

    consultas = Consulta.objects.filter(owner=request.user).select_related('paciente', 'dentista', 'procedimento').order_by('-data')

    # 🔍 Filtro por nome do paciente ou dentista
    if search:
        consultas = consultas.filter(
            Q(paciente__nome__icontains=search) |
            Q(dentista__nome__icontains=search)
        )

    # ✅ Filtro por status
    if status == 'pendente':
        consultas = consultas.filter(concluida=False)
    elif status == 'concluida':
        consultas = consultas.filter(concluida=True)

    # 📅 Filtro por data específica
    if data_filtro:
        try:
            data_formatada = parse_date(data_filtro)
            if data_formatada:
                consultas = consultas.filter(data__date=data_formatada)
        except:
            pass
        

    return render(request, 'clinic/consultas_list.html', {
         'consultas': consultas,
        'search': search or '',
        'status': status or '',
        'data_filtro': data_filtro or '',
    })


@login_required
@require_active_subscription
def consulta_create(request):
    if request.method == 'POST':
        form = ConsultaForm(request.POST, user=request.user)
        if form.is_valid():
            consulta = form.save(commit=False)
            consulta.owner = request.user

            # Aplica automaticamente o valor do procedimento
            if consulta.procedimento:
                consulta.valor = consulta.procedimento.valor_base

            consulta.save()
            messages.success(request, "Consulta criada com sucesso!")
            return redirect('clinic:consultas_list')
        else:
            messages.error(request, "Corrija os erros abaixo.")
    else:
        form = ConsultaForm(user=request.user)

    return render(request, 'clinic/consulta_form.html', {'form': form})


@login_required
@require_active_subscription
def consulta_update(request, pk):
    consulta = get_object_or_404(Consulta, pk=pk, owner=request.user)

    if request.method == 'POST':
        form = ConsultaForm(request.POST, instance=consulta, user=request.user)
        if form.is_valid():
            consulta = form.save(commit=False)

            if consulta.procedimento:
                consulta.valor = consulta.procedimento.valor_base

            consulta.save()
            form.save_m2m()
            
            # integração financeira
            if consulta.paga:
               if not Income.objects.filter(consulta=consulta).exists():
                   Income.objects.create(
                       owner = request.user,
                       origem = 'consulta',
                       consulta = consulta,
                       descricao = f'Consulta - {consulta.paciente.nome}',
                       valor = consulta.valor,
                       data = consulta.data.date(),
                       pago = True,
                    )
               
            messages.success(request, "Consulta atualizada!")
            return redirect('clinic:consultas_list')
        else:
            messages.error(request, "Corrija os erros abaixo.")
    else:
        form = ConsultaForm(instance=consulta, user=request.user)
        

    return render(request, 'clinic/consulta_form.html', {'form': form})


@login_required
@require_active_subscription
def consulta_delete(request, pk):
    consulta = get_object_or_404(Consulta, pk=pk, owner=request.user)
    
    if request.method == 'POST':
        consulta.delete()
        messages.success(request, "Consulta excluída com sucesso.")
        return redirect('clinic:consultas_list')
        
    return render(request, 'clinic/consulta_confirm_delete.html', {'consulta': consulta})

@login_required
@require_active_subscription
def consultas_calendar(request):
    # AJAX → retorna eventos
    if request.GET.get('start') and request.GET.get('end'):
        dentista_id = request.GET.get('dentista')

        consultas = Consulta.objects.filter(owner=request.user)

        if dentista_id:
            consultas = consultas.filter(dentista_id=dentista_id)

        eventos = []
        for c in consultas:
            eventos.append({
                "id": c.id,
                "title": c.paciente.nome,
                "start": c.data.isoformat(),
                "backgroundColor": "#0b5394" if not c.concluida else "#28a745",
                "borderColor": "#0b5394",
                "textColor": "white",
                "extendedProps": {
                    "dentista": c.dentista.nome if c.dentista else "",
                    "procedimento": c.procedimento.nome if c.procedimento else "",
                    "observacoes": c.observacoes or ""
                },
                "url": f"/consultas/{c.id}/editar/"
            })

        return JsonResponse(eventos, safe=False)

    # Tela normal
    return render(request, 'clinic/consultas_calendar.html', {
        'pacientes': Paciente.objects.filter(owner=request.user),
        'dentistas': Dentista.objects.filter(owner=request.user),
        'procedimentos': Procedimento.objects.filter(owner=request.user),
    })
    
@csrf_exempt
@login_required
@require_active_subscription
def consulta_update_ajax(request):
    """
    Atualiza a data/hora de uma consulta arrastada no calendário (FullCalendar).
    """
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método inválido"})

    consulta_id = request.POST.get("id")
    nova_data = request.POST.get("start")

    if not consulta_id or not nova_data:
        return JsonResponse({"success": False, "error": "Dados incompletos"})

    try:
        data_convertida = parse_datetime(nova_data)
        if not data_convertida:
            return JsonResponse({"success": False, "error": "Data inválida"})

        consulta = Consulta.objects.get(pk=consulta_id, owner=request.user)
        consulta.data = data_convertida
        consulta.save()

        return JsonResponse({"success": True})

    except Consulta.DoesNotExist:
        return JsonResponse({"success": False, "error": "Consulta não encontrada"})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
@require_active_subscription
def dashboard(request):
    user = request.user
    hoje = timezone.now().date()
    periodo = int(request.GET.get('periodo', 30))
    data_inicial = hoje - timedelta(days=periodo)

    # Garantir que o usuário tenha ao menos 1 dentista
    if not Dentista.objects.filter(owner=request.user).exists():
        return redirect('clinic:dentista_principal')

    # === CONSULTAS PAGAS do usuário ===
    consultas_pagas = Consulta.objects.filter(
        owner=user,
        paga=True,
        data__date__gte=data_inicial,
        data__date__lte=hoje + timedelta(days=90)
    )

    # Se não houver pagas no período → pega todas pagas do usuário
    if not consultas_pagas.exists():
        consultas_pagas = Consulta.objects.filter(owner=user, paga=True)

    # === Estatísticas Gerais ===
    total_pacientes = Paciente.objects.filter(owner=user).count()
    total_consultas = Consulta.objects.filter(owner=user).count()
    consultas_concluidas = consultas_pagas.filter(concluida=True).count()
    consultas_pendentes = Consulta.objects.filter(owner=user, concluida=False).count()

    # === Faturamento ===
    faturamento_total = consultas_pagas.aggregate(
        total=Sum('valor_final')
    )['total'] or 0

    faturamento_medio = consultas_pagas.aggregate(
        media=Avg('valor_final')
    )['media'] or 0

    comissoes_total = consultas_pagas.aggregate(
        total=Sum('comissao_valor')
    )['total'] or 0

    faturamento_liquido = faturamento_total - comissoes_total

    # === Faturamento mensal ===
    consultas_mes = consultas_pagas.filter(
        data__month=hoje.month,
        data__year=hoje.year
    )

    faturamento_mensal_bruto = consultas_mes.aggregate(
        total=Sum('valor_final')
    )['total'] or 0

    faturamento_mensal_comissoes = consultas_mes.aggregate(
        total=Sum('comissao_valor')
    )['total'] or 0

    faturamento_mensal_liquido = faturamento_mensal_bruto - faturamento_mensal_comissoes

    # === Estatísticas por dentista (somente do usuário) ===
    consultas_por_dentista = (
        consultas_pagas.exclude(dentista__isnull=True)
        .values('dentista__nome')
        .annotate(
            total_consultas=Count('id'),
            receita=Sum('valor_final'),
            comissao=Sum('comissao_valor')
        )
        .order_by('-receita')
    )

    if consultas_por_dentista:
        dentistas_labels = [c['dentista__nome'] for c in consultas_por_dentista]
        dentistas_qtd = [c['total_consultas'] for c in consultas_por_dentista]
        dentistas_receita = [float(c['receita'] or 0) for c in consultas_por_dentista]
        dentistas_comissao = [float(c['comissao'] or 0) for c in consultas_por_dentista]
    else:
        dentistas_labels = ['Sem dados']
        dentistas_qtd = [0]
        dentistas_receita = [0]
        dentistas_comissao = [0]

    # === Ranking ===
    ranking_dentistas = (
        consultas_pagas.values('dentista__nome')
        .annotate(
            total_consultas=Count('id'),
            receita=Sum('valor_final'),
            comissao=Sum('comissao_valor')
        )
        .order_by('-receita')[:5]
    )

    # === Últimos 6 meses ===
    meses = []
    dados_consultas = []
    dados_receita = []

    for i in range(5, -1, -1):
        mes_ref = hoje - timedelta(days=30 * i)
        nome_mes = calendar.month_abbr[mes_ref.month]

        cons_mes = consultas_pagas.filter(
            data__month=mes_ref.month,
            data__year=mes_ref.year
        )

        meses.append(nome_mes)
        dados_consultas.append(cons_mes.count())
        dados_receita.append(float(
            cons_mes.aggregate(total=Sum('valor_final'))['total'] or 0
        ))

    # === Status ===
    status_consultas = {
        'concluidas': consultas_concluidas,
        'pendentes': consultas_pendentes,
    }

    # === Próximas consultas (independente de pagas) ===
    inicio_semana = hoje
    fim_semana = hoje + timedelta(days=7)
    proximas_consultas = Consulta.objects.filter(
        owner=user,
        data__date__range=[inicio_semana, fim_semana]
    ).select_related('paciente', 'dentista').order_by('data')[:8]

    # === Assinatura ===
    assinatura = Assinatura.objects.filter(user=user).first()
    ultimo_pgto = (
        Pagamento.objects.filter(assinatura=assinatura, status='pago')
        .order_by('-data_pagamento')
        .first()
        if assinatura else None
    )

    plano_atual = None
    validade = None

    if assinatura:
        validade = assinatura.fim_teste

        if ultimo_pgto:
            plano_atual = ultimo_pgto.plano
        else:
            plano_atual = assinatura.tipo

    contexto = {
        # Geral
        'total_pacientes': total_pacientes,
        'total_consultas': total_consultas,
        'consultas_concluidas': consultas_concluidas,
        'consultas_pendentes': consultas_pendentes,

        # Gráficos
        'meses': meses,
        'dados_consultas': dados_consultas,
        'dados_receita': dados_receita,
        'dentistas_labels': dentistas_labels,
        'dentistas_qtd': dentistas_qtd,
        'dentistas_receita': dentistas_receita,
        'dentistas_comissao': dentistas_comissao,
        'ranking_dentistas': ranking_dentistas,

        # Financeiro
        'faturamento_total': faturamento_total,
        'faturamento_liquido': faturamento_liquido,
        'faturamento_mensal': faturamento_mensal_liquido,
        'faturamento_medio': faturamento_medio,
        'comissoes_total': comissoes_total,

        # Status / Outras informações
        'periodo': periodo,
        'status_consultas': status_consultas,
        'proximas_consultas': proximas_consultas,

        # Assinatura
        'plano_atual': plano_atual,
        'validade': validade
    }

    return render(request, 'clinic/dashboard.html', contexto)

@login_required
@require_active_subscription
def dashboard_data(request):
    user = request.user
    hoje = timezone.now().date()
    periodo = int(request.GET.get('periodo', 30))
    data_inicial = hoje - timedelta(days=periodo)

    # Consultas pagas do usuário
    consultas_pagas = Consulta.objects.filter(
        owner=user,
        paga=True,
        data__date__gte=data_inicial,
        data__date__lte=hoje + timedelta(days=90)
    )

    if not consultas_pagas.exists():
        consultas_pagas = Consulta.objects.filter(owner=user, paga=True)

    # Faturamentos
    faturamento_total = consultas_pagas.aggregate(
        total=Sum('valor_final')
    )['total'] or 0

    faturamento_medio = consultas_pagas.aggregate(
        media=Avg('valor_final')
    )['media'] or 0

    comissoes_total = consultas_pagas.aggregate(
        total=Sum('comissao_valor')
    )['total'] or 0

    faturamento_liquido = faturamento_total - comissoes_total

    # Mensal
    consultas_mes = consultas_pagas.filter(
        data__year=hoje.year,
        data__month=hoje.month
    )

    faturamento_mensal_bruto = consultas_mes.aggregate(
        total=Sum('valor_final')
    )['total'] or 0

    faturamento_mensal_comissoes = consultas_mes.aggregate(
        total=Sum('comissao_valor')
    )['total'] or 0

    faturamento_mensal_liquido = faturamento_mensal_bruto - faturamento_mensal_comissoes

    # Status
    status_consultas = {
        'concluidas': consultas_pagas.filter(concluida=True).count(),
        'pendentes': Consulta.objects.filter(owner=user, concluida=False).count(),
    }

    # Consultas por dentista
    consultas_por_dentista = (
        consultas_pagas.exclude(dentista__isnull=True)
        .values('dentista__nome')
        .annotate(
            total_consultas=Count('id'),
            receita=Sum('valor_final'),
            comissao=Sum('comissao_valor'),
        )
        .order_by('-receita')
    )

    # Gráfico 6 meses
    meses = []
    dados_consultas = []
    dados_receita = []

    for i in range(5, -1, -1):
        mes_ref = hoje - timedelta(days=30 * i)
        nome_mes = calendar.month_abbr[mes_ref.month]

        cons_mes = consultas_pagas.filter(
            data__month=mes_ref.month,
            data__year=mes_ref.year
        )

        meses.append(nome_mes)
        dados_consultas.append(cons_mes.count())
        dados_receita.append(float(
            cons_mes.aggregate(total=Sum('valor_final'))['total'] or 0
        ))

    return JsonResponse({
        'faturamento_total': float(faturamento_total),
        'faturamento_liquido': float(faturamento_liquido),
        'faturamento_mensal': float(faturamento_mensal_liquido),
        'faturamento_medio': float(faturamento_medio),
        'comissoes_total': float(comissoes_total),
        'status_consultas': status_consultas,
        'dentistas': list(consultas_por_dentista),
        'meses': meses,
        'dados_consultas': dados_consultas,
        'dados_receita': dados_receita,
    })


def _accent_insensitive_regex(prefix: str) -> str:
    """
    Converte 'brasilia' em uma regex que aceita acentos e casa no INÍCIO:
    '^br[aáàâã]sil[ií]a' etc. Mantém case-insensitive no __iregex.
    """
    base = {
        'a': '[aáàâã]', 'e': '[eéê]', 'i': '[ií]', 'o': '[oóôõ]', 'u': '[uú]',
        'c': '[cç]'
    }
    # escapa tudo e depois troca letras por classes
    esc = re.escape(prefix.strip().lower())
    # aplica mapeamento
    pattern = ''
    for ch in esc:
        pattern += base.get(ch, ch)
    return '^' + pattern  # começa com

@login_required
@require_active_subscription
def dentista_create(request):
    assinatura = Assinatura.objects.filter(user=request.user).first()
    
    if not assinatura:
        messages.error(request, 'Não foi possível encontrar sua assinatura')
        return redirect('clinic:dashboard')
    
    plano = assinatura.tipo

    # Limites por plano
    LIMITE = {
        "trial": 1,
        "basico": 1,
        "profissional": 4,
        "premium": 9999,
    }

    limite_max = LIMITE.get(plano, 1)
    
    dentistas_atual = Dentista.objects.filter(owner=request.user).count()

    # Bloqueia se atingiu o limite
    if dentistas_atual >= limite_max:
        messages.error(
            request,
            f"❌ Seu plano ({assinatura.get_tipo_display()}) permite no máximo {limite_max} dentista(s)."
        )
        return redirect("clinic:dentistas_list")

    from .forms import DentistaForm

    if request.method == "POST":
        form = DentistaForm(request.POST)
        if form.is_valid():
            dentista = form.save(commit=False)
            dentista.owner = request.user
            dentista.save()
            messages.success(request, "Dentista cadastrado com sucesso!")
            return redirect("clinic:dentistas_list")
        else:
            messages.error(request, 'Corrija os erros abaixo')
    else:
        form = DentistaForm()

    return render(request, "clinic/dentista_form.html", {"form": form})

@login_required
@require_active_subscription
def dentista_edit(request, id):
    dentista = get_object_or_404(Dentista, id=id, owner=request.user)

    from .forms import DentistaForm

    if request.method == "POST":
        form = DentistaForm(request.POST, instance=dentista)
        if form.is_valid():
            form.save()
            messages.success(request, "Dentista atualizado com sucesso!")
            return redirect("clinic:dentistas_list")
    else:
        form = DentistaForm(instance=dentista)

    return render(request, "clinic/dentista_form.html", {
        "form": form,
        "titulo": "Editar Dentista"
    })


@login_required
@require_active_subscription
def dentista_delete(request, id):
    dentista = get_object_or_404(Dentista, id=id, owner=request.user)
    dentista.delete()
    messages.success(request, "Dentista removido.")
    return redirect("clinic:dentistas_list")


@login_required
@require_active_subscription
def dentista_principal(request):
    """
    Tela especial para cadastrar o dentista principal,
    chamada no onboarding quando o usuário ainda não tem dentistas.
    """

    # 🔍 Se o usuário já tem dentista, manda pro dashboard
    if Dentista.objects.filter(owner=request.user).exists():
        return redirect("clinic:dashboard")

    if request.method == "POST":
        nome = request.POST.get("nome")
        cro = request.POST.get("cro")
        esp = request.POST.get("especialidade", "")
        tel = request.POST.get("telefone", "")
        email = request.POST.get("email", "")

        if not nome or not cro:
            messages.error(request, "Nome e CRO são obrigatórios.")
            return redirect("clinic:dentista_principal")

        Dentista.objects.create(
            owner=request.user,
            nome=nome,
            cro=cro,
            especialidade=esp,
            telefone=tel,
            email=email,
            comissao_percentual=0  # pode ser 0% por padrão e depois ajusta
        )

        messages.success(request, "Dentista principal cadastrado com sucesso!")
        return redirect("clinic:dashboard")

    return render(request, "clinic/dentista_principal.html", {
        "nome_sugerido": request.user.first_name or request.user.username
    })


@login_required
@require_active_subscription
def dentistas_list(request):
    assinatura = Assinatura.objects.filter(user=request.user).first()
    plano = assinatura.tipo if assinatura else 'trial'

    LIMITE = {
        "trial": 1,
        "basico": 1,
        "profissional": 4,
        "premium": 9999,
    }

    limite_max = LIMITE.get(plano, 1)

    dentistas = Dentista.objects.filter(owner=request.user)
    dentistas_count = dentistas.count()

    return render(
        request,
        "clinic/dentistas_list.html",
        {
            "dentistas": dentistas,
            "dentistas_count": dentistas_count,
            "limite_max": limite_max,
            "assinatura": assinatura,
        }
    )


@login_required
@require_active_subscription
def pacientes_list(request):
    search = (request.GET.get('search') or '').strip()
    pacientes = Paciente.objects.filter(owner=request.user).order_by('-data_cadastro')

    if search:
        regex = _accent_insensitive_regex(search)
        pacientes = pacientes.filter(
            Q(nome__iregex=regex) | 
            Q(cidade__iregex=regex) |
            Q(cpf__icontains=search)
        )

    context = {
        'pacientes': pacientes,
        'search': search,
    }
    return render(request, 'clinic/pacientes_list.html', context)


@login_required
@require_active_subscription
def paciente_create(request):
    if request.method == 'POST':
        form = PacienteForm(request.POST)
        if form.is_valid():
            paciente = form.save(commit=False)
            paciente.owner = request.user
            paciente.save()
            messages.success(request, "✅ Paciente cadastrado com sucesso!")
            return redirect('clinic:pacientes_list')
        else:
            messages.error(
                request, "⚠️ Corrija os erros abaixo antes de salvar.")
    else:
        form = PacienteForm()

    return render(request, 'clinic/paciente_form.html', {'form': form, 'titulo': 'Novo Paciente'})


@login_required
@require_active_subscription
def paciente_update(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk, owner=request.user)
    if request.method == 'POST':
        form = PacienteForm(request.POST, instance=paciente)
        if form.is_valid():
            form.save()
            messages.success(request, "Dados do paciente atualizados!")
            return redirect('clinic:pacientes_list')
    else:
        form = PacienteForm(instance=paciente)
        
    return render(request, 'clinic/paciente_form.html', {'form': form, 'titulo': 'Editar Paciente'})


@login_required
@require_active_subscription
def paciente_delete(request, pk):
    paciente = get_object_or_404(Paciente, pk=pk, owner=request.user)
    if request.method == 'POST':
        paciente.delete()
        messages.success(request, "Paciente excluído com sucesso.")
        return redirect('clinic:pacientes_list')
    return render(request, 'clinic/paciente_confirm_delete.html', {'paciente': paciente})


@csrf_exempt
@login_required
@require_active_subscription
def consulta_create_ajax(request):
    if request.method == "POST":
        try:
            paciente_id = request.POST.get("paciente")
            dentista_id = request.POST.get("dentista")
            procedimento_id = request.POST.get("procedimento")
            data = request.POST.get("data")
            observacoes = request.POST.get("observacoes", "")

            consulta = Consulta.objects.create(
                paciente_id=paciente_id,
                dentista_id=dentista_id,
                procedimento_id=procedimento_id,
                data=data,
                observacoes=observacoes,
            )
            return JsonResponse({"success": True, "id": consulta.id})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método inválido"})


def registrar_teste(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']

        # Verifica se o nome de usuário já existe
        if User.objects.filter(username=username).exists():
            messages.error(request, "Usuário já existe.")
            return redirect('clinic:registrar_teste')

        # Verifica se o e-mail foi informado
        if not email:
            messages.error(
                request, "O e-mail é obrigatório para criar uma conta.")
            # 👈 aqui ajustei o namespace
            return redirect('clinic:registrar_teste')

        # Cria o usuário
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )
        user.date_joined = timezone.now()
        user.save()
        
        # Cria assinatura trial automaticamente
        Assinatura.objects.create(user=user, tipo='trial')

        # Envia email de boas vindas
        nome = user.first_name or user.username

        send_mail(
            subject="🎉 Bem-vindo ao OdontoIA!",
            message=f"Olá {nome}, seja bem-vindo ao OdontoIA!\nSua conta de teste está ativa por 7 dias.",
            from_email=None,  # usa o DEFAULT_FROM_EMAIL
            recipient_list=[user.email],
            fail_silently=True,
        )
        
        # Faz login automático após criar conta
        login(request, user)

        messages.success(
            request,
            f"Conta de teste criada com sucesso! Bem-vindo(a), {user.username}.Aproveite seus 7 dias gratuitos."
        )
        return redirect('clinic:onboarding')

    # GET → renderiza o formulário normalmente
    return render(request, 'clinic/registrar_teste.html')


@login_required
@require_active_subscription
def procedimentos_list(request):
    procedimentos = Procedimento.objects.filter(owner=request.user).order_by('nome')
    return render(request, 'clinic/procedimentos_list.html', {'procedimentos': procedimentos})


def procedimento_valor(request, id):
    procedimento = Procedimento.objects.get(id=id)
    return JsonResponse({'valor': procedimento_valor})


@login_required
@require_active_subscription
def procedimento_create(request):
    if request.method == 'POST':
        form = ProcedimentoForm(request.POST)
        
        if form.is_valid():
            procedimento = form.save(commit=False)
            procedimento.owner = request.user
            procedimento.save()
            messages.success(request, 'Procedimento criado com sucesso!')
            return redirect('clinic:procedimentos_list')
        else:
            messages.error(request, 'Corrija os erros antes de salvar')
    else:        
        form = ProcedimentoForm()
        
    return render(request, 'clinic/procedimento_form.html', {'form': form, 'titulo': 'Novo Procedimento'})


@login_required
@require_active_subscription
def procedimento_edit(request, id):
    procedimento = get_object_or_404(Procedimento, id=id, owner=request.user)
    
    if request.method == 'POST':
        form = ProcedimentoForm(request.POST, instance=procedimento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Procedimento atualizado com sucesso!')
            return redirect('clinic:procedimentos_list')
        else:
            messages.error(request, 'Corrija os erros abaixo.')
    else:
        form = ProcedimentoForm(instance=procedimento)
        
    return render(request, 'clinic/procedimento_form.html', {'form': form, 'titulo': 'Editar Procedimento'})


@login_required
@require_active_subscription
def procedimento_delete(request, id):
    procedimento = get_object_or_404(Procedimento, id=id, owner=request.user)
    
    if request.method == 'POST':
        procedimento.delete()
        messages.success(request, "Procedimento excluído com sucesso.")
        return redirect('clinic:procedimentos_list')
    
    return render(request, 'clinic/procedimento_confirm_delete.html', {'procedimento': procedimento})


def assinatura_expirada(request):
    return render(request, "clinic/assinatura_expirada.html")


@csrf_exempt
def password_reset_request(request):
    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        if not email:
            messages.error(request, "Por favor, informe um e-mail válido.")
            return redirect("login")

        # Verifica se há um usuário com este e-mail
        user = User.objects.filter(email=email).first()
        if not user:
            messages.warning(
                request, "E-mail não encontrado. Verifique ou entre em contato com o suporte.")
            return redirect("login")

        # Simulação de envio de e-mail (você pode configurar SMTP real)
        assunto = "Recuperação de acesso - OdontoIA"
        corpo = f"""
Olá, {user.username}!

Recebemos uma solicitação para recuperar o acesso ao OdontoIA.

Usuário: {user.username}
Se esqueceu sua senha, redefina acessando: https://odontoia.codertec.com.br/resetar-senha/

Atenciosamente,
Equipe OdontoIA
        """
        try:
            send_mail(
                assunto,
                corpo,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            messages.success(
                request, "E-mail enviado com sucesso! Verifique sua caixa de entrada.")
        except Exception as e:
            print("Erro ao enviar e-mail:", e)
            messages.error(
                request, "Não foi possível enviar o e-mail. Tente novamente mais tarde.")

        return redirect("login")

    return redirect("login")


# ===============================
# 💌 Recuperação de Senha (HTML)
# ===============================
@csrf_exempt
def password_reset_request(request):
    """
    Exibe modal e envia e-mail estilizado com link de redefinição de senha.
    """
    if request.method == "POST":
        email = request.POST.get("email")
        if not email:
            return JsonResponse({"success": False, "error": "Informe o e-mail de cadastro."})

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return JsonResponse({"success": False, "error": "E-mail não encontrado no sistema."})

        # Gera token de redefinição
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes

        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        reset_url = request.build_absolute_uri(f"/reset/{uid}/{token}/")

        # Renderiza template HTML do e-mail
        context = {
            "user": user,
            "reset_url": reset_url,
        }
        html_content = render_to_string(
            "clinic/emails/password_reset_email.html", context)
        subject = "🔑 Redefinição de senha - OdontoIA"

        msg = EmailMultiAlternatives(
            subject,
            f"Olá {user.username},\n\nPara redefinir sua senha, acesse o link abaixo:\n{reset_url}\n\nEquipe OdontoIA",
            "OdontoIA <no-reply@odontoia.com.br>",
            [email],
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send(fail_silently=False)

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Método inválido."})


@login_required
def onboarding(request):
    """
    Mostra uma tela de boas-vindas e primeiros passos para novos usuários.
    Só aparece uma vez por usuário.
    """
    if request.session.get('onboarding_done'):
        return redirect('clinic:dashboard')

    assinatura = Assinatura.objects.filter(user=request.user).first()
    tipo_assinatura = assinatura.tipo if assinatura else 'trial'

    if request.method == 'POST':
        request.session['onboarding_done'] = True
        return redirect('clinic:dentista_principal')

    return render(request, 'clinic/onboarding.html')


# ==== PAGAMENTOS (Mercado Pago) ============================================

def _get_mp_sdk():
    """
    Retorna o SDK autenticado do Mercado Pago, ou lança erro amigável se faltar token.
    """
    import mercadopago
    from django.conf import settings

    access_token = getattr(settings, "MERCADOPAGO_ACCESS_TOKEN", None)
    if not access_token:
        print("❌ Falha: variável MERCADOPAGO_ACCESS_TOKEN ausente no settings.")
        raise RuntimeError("MERCADOPAGO_ACCESS_TOKEN não configurado.")

    try:
        return mercadopago.SDK(access_token)
    except Exception as e:
        print(f"❌ Erro ao inicializar SDK do Mercado Pago: {e}")
        raise


PLANOS = {
    "basico": Decimal("49.90"),
    "profissional": Decimal("79.90"),
    "premium": Decimal("129.90"),
}


@login_required
def criar_pagamento(request, plano: str):

    plano = plano.lower().strip()
    PLANOS = {
        "basico": Decimal("49.90"),
        "profissional": Decimal("79.90"),
        "premium": Decimal("129.90"),
    }

    if plano not in PLANOS:
        messages.error(request, "Plano inválido.")
        return redirect("clinic:dashboard")

    valor = float(PLANOS[plano])

    # Assinatura do usuário
    assinatura, _ = Assinatura.objects.get_or_create(user=request.user)

    # Registrar pagamento
    referencia = f"odontoia-{request.user.id}-{uuid.uuid4().hex}"
    pagamento = Pagamento.objects.create(
        assinatura=assinatura,
        referencia=referencia,
        valor=valor,
        status="pendente",
        metodo="desconhecido",
    )

    # Mercado Pago
    access_token = getattr(settings, "MERCADOPAGO_ACCESS_TOKEN", None)
    if not access_token:
        messages.error(request, "Token Mercado Pago não configurado.")
        return redirect("clinic:dashboard")

    sdk = mercadopago.SDK(access_token)

    base_url = (
        "https://app.odontoia.codertec.com.br"
        if not settings.DEBUG
        else request.build_absolute_uri("/")[:-1]
    )

    preference_data = {
        "items": [
            {
                "id": referencia,
                "title": f"Plano {plano.capitalize()} - OdontoIA",
                "quantity": 1,
                "currency_id": "BRL",
                "unit_price": valor,
            }
        ],
        "payer": {"email": request.user.email},
        "back_urls": {
            "success": f"{base_url}/pagamento/sucesso/",
            "failure": f"{base_url}/pagamento/falha/",
            "pending": f"{base_url}/pagamento/sucesso/",
        },
        "external_reference": referencia,
    }

    if not settings.DEBUG:
        preference_data["notification_url"] = f"{base_url}/mp/webhook/"
        preference_data["auto_return"] = "approved"

    pref = sdk.preference().create(preference_data)
    resp = pref.get("response", {})
    init_point = resp.get("init_point") or resp.get("sandbox_init_point")

    if not init_point:
        messages.error(request, "Erro ao iniciar pagamento.")
        return redirect("clinic:dashboard")

    return redirect(init_point)


# ===========================
# 🧾 WEBHOOK DO MERCADO PAGO
# ===========================
@csrf_exempt
def mercadopago_webhook(request):
    """
    Recebe notificações automáticas do Mercado Pago sobre pagamentos.
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    payment_id = payload.get("data", {}).get("id") or payload.get("id")
    if not payment_id:
        return JsonResponse({"ok": True, "ignored": True})

    sdk = _get_mp_sdk()
    payment_info = sdk.payment().get(payment_id).get("response", {})

    status = payment_info.get("status")
    external_reference = payment_info.get("external_reference")
    payment_method = (payment_info.get("payment_method_id") or "desconhecido").lower()

    if not external_reference:
        return JsonResponse({"ok": True, "missing_external_reference": True})

    pgto = Pagamento.objects.filter(referencia=external_reference).first()
    if not pgto:
        return JsonResponse({"ok": True, "unknown_reference": True})

    pgto.raw_payload = payload
    pgto.metodo = (
        "pix" if "pix" in payment_method
        else "card" if any(k in payment_method for k in ["visa", "master", "amex", "hiper", "elo"])
        else "boleto" if "boleto" in payment_method
        else "desconhecido"
    )

    if status == "approved":
        pgto.status = "pago"
        pgto.data_pagamento = timezone.now()
        pgto.save()

        assinatura = pgto.assinatura
        assinatura.ativa = True
        assinatura.fim_teste = timezone.now() + timezone.timedelta(days=30)
        assinatura.save()
    elif status in ("rejected", "cancelled", "refunded", "charged_back"):
        pgto.status = "falhou"
        pgto.save()
    else:
        pgto.status = "pendente"
        pgto.save()

    return JsonResponse({"ok": True})


# ===========================
# ✅ PÁGINA DE SUCESSO
# ===========================
@login_required
def pagamento_sucesso(request):
    assinatura = Assinatura.objects.filter(user=request.user).first()
    pagamento = (
        Pagamento.objects.filter(assinatura__user=request.user)
        .order_by("-data_pagamento")
        .first()
    )

    if not assinatura or not pagamento:
        messages.warning(request, "Não foi possível confirmar o pagamento.")
        return redirect("clinic:dashboard")

    # Ativa assinatura por 30 dias
    assinatura.ativa = True
    assinatura.fim_teste = timezone.now() + timedelta(days=30)
    assinatura.save()

    # ---- Email de confirmação (somente para assinatura paga) ----
    from django.template.loader import render_to_string

    nome = request.user.first_name or request.user.username
    plano = pagamento.plano.capitalize()
    validade = assinatura.fim_teste.strftime("%d/%m/%Y")

    html_email = render_to_string(
        "clinic/emails/assinatura_ativada.html",
        {"nome": nome, "plano": plano, "validade": validade}
    )

    msg = EmailMultiAlternatives(
        subject="🎉 Assinatura ativada - OdontoIA",
        body=f"Sua assinatura {plano} está ativa até {validade}.",
        from_email="OdontoIA <no-reply@odontoia.com.br>",
        to=[request.user.email],
    )
    msg.attach_alternative(html_email, "text/html")
    msg.send(fail_silently=True)
    # --------------------------------------------------------------

    return render(request, "clinic/pagamento_sucesso.html", {
        "assinatura": assinatura,
        "pagamento": pagamento,
        "plano": plano,
    })

    

# Checkout publico
def checkout_publico(request, plano):
    plano = plano.lower()

    PLANOS = {
        "basico": 49.90,
        "profissional": 79.90,
        "premium": 129.90,
    }

    if plano not in PLANOS:
        messages.error(request, "Plano inválido.")
        return redirect("https://odontoia.codertec.com.br")

    valor = PLANOS[plano]

    # GET → mostra o formulário
    if request.method == "GET":
        return render(request, "clinic/checkout_publico.html", {
            "plano_nome": plano.capitalize(),
            "plano_slug": plano,
            "valor": valor
        })

    # POST → o usuário enviou nome e email, vamos criar uma conta trial
    nome = request.POST.get("nome")
    email = request.POST.get("email")

    if not email:
        messages.error(request, "E-mail é obrigatório.")
        return redirect(request.path)

    # Cria usuário temporário (ou pega existente)
    user, created = User.objects.get_or_create(
        username=email,
        defaults={"email": email, "first_name": nome or email}
    )

    if created:
        # cria assinatura trial automática
        Assinatura.objects.create(user=user, tipo=plano)

    # Faz login automático
    login(request, user)

    # Redireciona para o checkout interno
    return redirect(f"/pagamento/checkout/{plano}/")


# ===========================
# ❌ PÁGINA DE FALHA
# ===========================
@login_required
def pagamento_falha(request):
    pagamento = (
        Pagamento.objects.filter(assinatura__user=request.user)
        .order_by("-data_pagamento")
        .first()
    )

    plano = pagamento.plano if pagamento else "Indefinido"

    return render(request, "clinic/pagamento_falha.html", {"plano": plano})


# --- FINANCEIRO (apenas Profissional / Premium) ---------------------------
@login_required
@require_active_subscription
def financeiro_home(request):
    # Permissão: somente profissional ou premium
    assinatura = Assinatura.objects.filter(user=request.user).first()

    if not assinatura or assinatura.tipo not in ["profissional", "premium"]:
        messages.error(request, "Seu plano não permite acesso ao Financeiro.")
        return redirect("clinic:dashboard")

    return render(request, "clinic/financeiro_home.html")


@login_required
@require_active_subscription
def financeiro_resumo(request):
    """
    Tela de resumo financeiro (por dentista, receita, comissões, líquido)
    Disponível apenas para planos Profissional e Premium.
    """
    assinatura = Assinatura.objects.filter(user=request.user, ativa=True).first()

    if not assinatura or assinatura.tipo not in ("profissional", "premium"):
        messages.error(
            request,
            "O módulo financeiro completo está disponível apenas para os planos "
            "Profissional e Premium."
        )
        return redirect("clinic:dashboard")

    hoje = timezone.now().date()
    periodo = int(request.GET.get("periodo", 30))  # dias
    data_inicial = hoje - timedelta(days=periodo)

    consultas = Consulta.objects.filter(
        owner=request.user,
        data__date__gte=data_inicial,
        data__date__lte=hoje
    )

    # Agrupado por dentista
    por_dentista = (
        consultas.exclude(dentista__isnull=True)
        .values("dentista__nome")
        .annotate(
            total_consultas=Count("id"),
            receita=Sum("valor_final"),
            comissoes=Sum("comissao_valor"),
        )
        .order_by("-receita")
    )
    
    # Cálculo líquido por dentista
    lista_dentistas = []
    for d in por_dentista:
        receita = d['receita'] or 0
        comissoes = d['comissoes'] or 0
        d['liquido'] = receita - comissoes 
        lista_dentistas.append(d)
        
    totais = consultas.aggregate(
        total_receita=Sum("valor_final"),
        total_comissoes=Sum("comissao_valor"),
    )
    total_receita = totais["total_receita"] or 0
    total_comissoes = totais["total_comissoes"] or 0
    total_liquido = total_receita - total_comissoes
    
    # Adiciona valor líquido por dentista
    lista_dentistas = []
    for dent in por_dentista:
        receita = dent['receita'] or 0
        comissao = dent['comissoes'] or 0
        dent['liquido'] = receita - comissao 
        lista_dentistas.append(dent)

    context = {
        "por_dentista": lista_dentistas,
        "periodo": periodo,
        "total_receita": total_receita,
        "total_comissoes": total_comissoes,
        "total_liquido": total_liquido,
    }
    
    
    return render(request, "clinic/financeiro_resumo.html", context)


@login_required
@require_active_subscription
def financeiro_exportar_excel(request):
    """
    Exporta o mesmo resumo financeiro para Excel (XLSX).
    Também restrito a Profissional / Premium.
    """
    assinatura = Assinatura.objects.filter(user=request.user, ativa=True).first()

    if not assinatura or assinatura.tipo not in ("profissional", "premium"):
        messages.error(
            request,
            "Exportação disponível apenas para os planos Profissional e Premium."
        )
        return redirect("clinic:dashboard")

    hoje = timezone.now().date()
    periodo = int(request.GET.get("periodo", 30))
    data_inicial = hoje - timedelta(days=periodo)

    consultas = Consulta.objects.filter(
        owner = request.user,
        data__date__gte=data_inicial,
        data__date__lte=hoje
    )

    por_dentista = (
        consultas.exclude(dentista__isnull=True)
        .values("dentista__nome")
        .annotate(
            total_consultas=Count("id"),
            receita=Sum("valor_final"),
            comissoes=Sum("comissao_valor"),
        )
        .order_by("-receita")
    )

    # --- Monta planilha Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financeiro"

    headers = ["Dentista", "Total Consultas", "Receita (R$)", "Comissões (R$)", "Líquido (R$)"]
    ws.append(headers)

    for row in por_dentista:
        receita = row["receita"] or 0
        comissoes = row["comissoes"] or 0
        liquido = receita - comissoes
        ws.append([
            row["dentista__nome"],
            row["total_consultas"],
            float(receita),
            float(comissoes),
            float(liquido),
        ])

    # Ajusta largura das colunas
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Resposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"financeiro_{hoje.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# IA e Insights (apenas profissional e premium)
@login_required
@require_active_subscription
def ia_insights(request):
    # Permissão: apenas Premium
    assinatura = Assinatura.objects.filter(user=request.user).first()

    if not assinatura or assinatura.tipo != "premium":
        messages.error(request, "Apenas assinantes Premium podem acessar IA & Insights.")
        return redirect("clinic:dashboard")

    return render(request, "clinic/ia_insights.html")


# Dashboard Financeiro
@login_required
@require_active_subscription
def financeiro_dashboard(request):
    from .services import get_fluxo_caixa
    
    # Filtros opcionais
    mes = request.GET.get("mes")
    ano = request.GET.get("ano")

    stats = get_fluxo_caixa(request.user, mes=mes, ano=ano)

    # Puxa lista de receitas e despesas para exibir no dashboard
    incomes = Income.objects.filter(owner=request.user).order_by('-data')[:10]
    expenses = Expense.objects.filter(owner=request.user).order_by('-data')[:10]

    context = {
        'stats': stats,
        'incomes': incomes,
        'expenses': expenses,
        'mes': mes,
        'ano': ano,
    }

    return render(request, 'clinic/financeiro_dashboard.html', context)


# CRUD Receitas e Despesas
@login_required
@require_active_subscription
def receitas_list(request):
    receitas = Income.objects.filter(owner = request.user).order_by('-data')
    return render(request, 'clinic/receitas_list.html', {'receitas': receitas})


@login_required
@require_active_subscription
def receita_create(request):
    if request.method == "POST":
        form = IncomeForm(request.POST)
        if form.is_valid():
            receita = form.save(commit=False)
            receita.owner = request.user
            receita.origem = 'manual'
            receita.save()
            messages.success(request, 'Receita adicionada!')
            return redirect('clinic:receitas_list')

    else:
        form = IncomeForm()
            
    return render(request, "clinic/receita_form.html", {'form': form})


@login_required
@require_active_subscription
def receita_update(request, pk):
    receita = get_object_or_404(Income, pk=pk, owner=request.user)

    if request.method == "POST":
        form = IncomeForm(request.POST, instance=receita)
        if form.is_valid():
            form.save()
            messages.success(request, "Receita atualizada!")
            return redirect("clinic:receitas_list")
    else:
        form = IncomeForm(instance=receita)

    return render(request, "clinic/receita_form.html", {"receita": receita})


@login_required
@require_active_subscription
def receita_delete(request, pk):
    receita = get_object_or_404(Income, pk=pk, owner=request.user)
    receita.delete()
    messages.success(request, "Receita removida!")
    return redirect("clinic:receitas_list")


# DESPESAS
@login_required
@require_active_subscription
def despesas_list(request):
    despesas = Expense.objects.filter(owner=request.user).order_by('-data')
    return render(request, "clinic/despesas_list.html", {"despesas": despesas})


@login_required
@require_active_subscription
def despesa_create(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.owner = request.user
            despesa.save()
            
            messages.success(request, "Despesa adicionada!")
            return redirect("clinic:despesas_list")
    else:
        form = ExpenseForm()

    return render(request, "clinic/despesa_form.html", {"form": form})


@login_required
@require_active_subscription
def despesa_update(request, pk):
    despesa = get_object_or_404(Expense, pk=pk, owner=request.user)

    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=despesa)
        if form.is_valid():
            form.save()
            messages.success(request, "Despesa atualizada!")
            return redirect("clinic:despesas_list")
    else:
        form = ExpenseForm(instance=despesa)

    return render(request, "clinic/despesa_form.html", {"form": form})


@login_required
@require_active_subscription
def despesa_delete(request, pk):
    despesa = get_object_or_404(Expense, pk=pk, owner=request.user)
    despesa.delete()
    messages.success(request, "Despesa removida!")
    return redirect("clinic:despesas_list")
